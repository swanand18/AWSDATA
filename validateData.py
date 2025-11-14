# validate_input.py

import pandas as pd
from sqlalchemy import text
from database import get_db
from functions import get_existing_values, validate_column, check_company_existence, check_contact_existence
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Path to input and output
CSV_PATH = "D:/Repositories/FinaFunnel/FinalFunnel-data/import-data.csv"
OUTPUT_PATH = "D:/Repositories/FinaFunnel/FinalFunnel-data/validated_output.csv"

# Columns to validate and corresponding dimension tables
VALIDATION_MAP = {
    "jobtitle": "dim_jobtitles",
    "managementlevel": "dim_manlevels",
    "emailstatus": "dim_emailstatuses",
    "country": "dim_countries",
    "compstate": "dim_states",
    "city": "dim_cities",
    "postalcode": "dim_postalcodes",
    "address": "dim_addresses",
    "industry": "dim_industries"
}

def main():
    df = pd.read_csv(CSV_PATH)

    with next(get_db()) as db:
        for column, dim_table in VALIDATION_MAP.items():
            if column in df.columns:
                print(f"Validating '{column}' against '{dim_table}'...")
                valid_values = get_existing_values(dim_table, db)
                df = validate_column(df, column, valid_values)
            else:
                print(f"⚠️ Column '{column}' not found in the file. Skipping.")

        if "companyname" in df.columns and "comp_domain" in df.columns:
            print("Checking company (name + domain) combination...")
            df = check_company_existence(df, db)

        if "emplinkedin" in df.columns and "empemail" in df.columns:
            print("Checking contact (LinkedIn + Email) combination...")
            df = check_contact_existence(df, db)

    # Save to Excel
    EXCEL_PATH = OUTPUT_PATH.replace(".csv", ".xlsx")
    df.to_excel(EXCEL_PATH, index=False)

    # Load workbook and apply formatting
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Define light green fill
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Apply fill to all *_status columns
    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name.endswith("_status"):
            for row in range(2, ws.max_row + 1):  # skip header
                cell = ws[f"{get_column_letter(col_idx)}{row}"]
                cell.fill = green_fill

    wb.save(EXCEL_PATH)
    print(f"✅ Excel file with formatting saved to {EXCEL_PATH}")

if __name__ == "__main__":
    main()
