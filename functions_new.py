# functions.py

import pandas as pd
import sys
import streamlit as st
from sqlalchemy import text
from app_backend.database import engine
import json
from app_backend.database import get_db
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import re
from datetime import datetime
import subprocess
import os

# Used for validation logic
VALIDATION_MAP = {
    "jobtitle": "dim_jobtitles",
    "managementlevel": "dim_manlevels",
    "emailstatus": "dim_emailstatuses",
    "country": "dim_countries",
    "compstate": "dim_states",
    "city": "dim_cities",
    "postalcode": "dim_postalcodes",
    "address": "dim_addresses",
    "industry": "dim_industries"
}

def get_filter_options(column):
    key = f"filter_options_{column}"
    if key not in st.session_state:
        VALID_COLUMNS = {
            "country", "city", "compstate", "industry", "emailstatus", "managementlevel",
            "jobtitle", "companyname", "comp_domain", "postalcode"
        }
        if column not in VALID_COLUMNS:
            raise ValueError(f"Invalid column name: {column}")

        query = f'SELECT DISTINCT "{column}" FROM cached_full_contacts_data ORDER BY "{column}"'
        with engine.connect() as conn:
            df = pd.read_sql(query, conn)
            st.session_state[key] = ["All"] + df[column].dropna().astype(str).tolist()

    return st.session_state[key]

#@st.cache_data(ttl=86400)
# def get_filter_options(column):
#     VALID_COLUMNS = {
#         "country", "city", "compstate", "industry", "emailstatus", "managementlevel",
#         "jobtitle", "companyname", "comp_domain", "postalcode"
#     }
#     if column not in VALID_COLUMNS:
#         raise ValueError(f"Invalid column name: {column}")

#     query = f'SELECT DISTINCT "{column}" FROM cached_full_contacts_data ORDER BY "{column}"'
#     with engine.connect() as conn:
#         df = pd.read_sql(query, conn)
#         return ["All"] + df[column].dropna().astype(str).tolist()

def save_query(query_name, filter_values):
    st.session_state.saved_queries[query_name] = filter_values
    st.success(f"Query '{query_name}' saved successfully!")

def load_query(query_name):
    if query_name in st.session_state.saved_queries:
        return st.session_state.saved_queries[query_name]
    else:
        st.warning(f"Query '{query_name}' not found!")
        return {}

def get_uploaded_filter_conditions(file, mode):
    conditions = []
    params = {}
    if file:
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            return [], {}

        if df.empty:
            return [], {}

        filter_column = df.columns[0]
        if filter_column not in ["Company Name", "Domain Name", "Zip Code"]:
            return [], {}

        column_map = {
            "Company Name": "CompanyName",
            "Domain Name": "comp_domain",
            "Zip Code": "PostalCode"
        }
        sql_column = column_map[filter_column]
        values = df[filter_column].dropna().astype(str).unique()

        placeholders = []
        for i, val in enumerate(values):
            key = f"{mode}_{i}"
            placeholders.append(f":{key}")
            params[key] = val

        if mode == "exclude":
            conditions.append(f"{sql_column} NOT IN ({', '.join(placeholders)})")
        elif mode == "include":
            conditions.append(f"{sql_column} IN ({', '.join(placeholders)})")

    return conditions, params

# Function to load saved queries
def load_saved_queries():
    db_session = next(get_db())
    query = text("SELECT id, name FROM dim_savedqueries ORDER BY timestamp DESC")  # Load id and name
    result = db_session.execute(query)
    df = pd.DataFrame(result.fetchall(), columns=result.keys())
    return df

# Function to load filters for a selected query
def load_query_filters(query_id):
    db_session = next(get_db())
    query = text("SELECT filters, name FROM dim_savedqueries WHERE id = :query_id")
    result = db_session.execute(query, {"query_id": int(query_id)}).fetchone()

    if result:
        filters = result[0]  # Get filters from the result
        campaign_name = result[1]  # Get campaign name

        if filters:  # If filters are present in the database
            try:
                # Parse filters if it's a JSON string
                filters = json.loads(filters)
                return filters, campaign_name
            except json.JSONDecodeError:
                st.error("Error decoding JSON from the saved query filters.")
                return None, campaign_name
        else:
            # If filters are empty, return None and the campaign name
            return None, campaign_name
    else:
        st.error("No filters or campaign name found for the selected query.")
        return None, None

def get_or_create_dim_ids(df, db_session, dim_table, column, id_column, create_missing=False):
    """
    Generic function to retrieve or insert dimension IDs.

    Parameters:
    - df: DataFrame containing the values to look up
    - db_session: Active SQLAlchemy DB session
    - dim_table: Table to query (e.g., 'dim_manlevels')
    - column: Column in df to look up
    - id_column: New column name for the resolved ID
    - create_missing: Whether to insert missing values into the dimension table

    Returns:
    - List of resolved IDs (same length as df)
    """
    records = db_session.execute(text(f"SELECT id, name FROM {dim_table}")).fetchall()
    value_map = {row.name.strip().lower(): row.id for row in records}
    result_ids = []

    for val in df[column]:
        if pd.isna(val) or str(val).strip() == "":
            result_ids.append(None)
            continue

        val_clean = str(val).strip().lower()
        existing_id = value_map.get(val_clean)

        if existing_id:
            result_ids.append(existing_id)
        elif create_missing:
            insert_stmt = text(f"INSERT INTO {dim_table}(name) VALUES (:name) RETURNING id")
            new_id = db_session.execute(insert_stmt, {"name": val.strip()}).scalar()
            db_session.commit()
            value_map[val_clean] = new_id
            result_ids.append(new_id)
        else:
            result_ids.append(None)

    return result_ids

def resolve_dim_ids(df_source, db_session, dim_table, column_name, id_column_name):
    """
    Matches values from a column in df_source to IDs in a dimension table.

    Returns a DataFrame with 'index' and the resolved ID column.
    """
    records = db_session.execute(text(f"SELECT id, name FROM {dim_table}")).fetchall()
    dim_map = {row.name.strip().lower(): row.id for row in records}

    resolved_ids = [
        dim_map.get(str(val).strip().lower(), None)
        if pd.notna(val) and str(val).strip() != "" else None
        for val in df_source[column_name]
    ]

    return pd.DataFrame({
        "index": df_source["index"],
        id_column_name: resolved_ids
    })

def get_existing_company_ids(df, db_session):
    """
    Checks if company already exists in the database using comp_domain first, then companyname.
    Returns a tuple of (matched_ids_dict, unmatched_rows_df)
    """
    matched_company_ids = {}
    unmatched_rows = []

    for _, row in df.iterrows():
        idx = row['index']
        domain = str(row['comp_domain']).strip().lower() if pd.notna(row['comp_domain']) else None
        name = str(row['companyname']).strip().lower() if pd.notna(row['companyname']) else None

        company_id = None

        if domain:
            result = db_session.execute(
                text("SELECT id FROM fact_companies WHERE LOWER(comp_domain) = :domain LIMIT 1"),
                {"domain": domain}
            ).fetchone()
            if result:
                company_id = result[0]

        if not company_id and name:
            result = db_session.execute(
                text("SELECT id FROM fact_companies WHERE LOWER(name) = :name LIMIT 1"),
                {"name": name}
            ).fetchone()
            if result:
                company_id = result[0]

        if company_id:
            matched_company_ids[idx] = company_id
        else:
            unmatched_rows.append(row)

    return matched_company_ids, pd.DataFrame(unmatched_rows)

def prepare_unique_companies(df_unmatched_companies):
    df_unique = df_unmatched_companies[
        ['index', 'companyname', 'comp_domain', 'comp_phone', 'comp_linkedin',
         'annrev', 'empsize', 'address', 'city', 'country', 'compstate', 'postalcode', 'industry']
    ].copy()
    df_unique.rename(columns={'companyname': 'name'}, inplace=True)
    df_unique = df_unique.drop_duplicates(subset=['comp_domain', 'name'])
    # DO NOT reset index — we keep original contact linkage
    return df_unique

def load_new_data(file_path):
    try:
        _, ext = os.path.splitext(file_path.lower())
        if ext in ['.xls', '.xlsx']:
            return pd.read_excel(file_path)
        elif ext == '.csv':
            return pd.read_csv(file_path)
        else:
            print(f"Unsupported file format: {ext}")
            return None
    except Exception as e:
        print(f"Error loading data: {e}")
        return None

def get_or_create_jobtitle_ids(df, db_session):
    """
    Resolves jobtitle IDs based on both jobtitle name and manlevel_id.
    Inserts new records if not found.
    """
    # Load existing jobtitles with manlevel_id
    existing = db_session.execute(text("SELECT id, name, manlevel_id FROM dim_jobtitles")).fetchall()
    jobtitle_map = {(row.name.strip().lower(), row.manlevel_id): row.id for row in existing}

    result_ids = []

    for _, row in df.iterrows():
        title = str(row['jobtitle']).strip() if pd.notna(row['jobtitle']) else ""
        manlevel_id = row['manlevel_id']

        if not title or pd.isna(manlevel_id):
            result_ids.append(None)
            continue

        key = (title.lower(), manlevel_id)
        jobtitle_id = jobtitle_map.get(key)

        if jobtitle_id:
            result_ids.append(jobtitle_id)
        else:
            # Insert new jobtitle with manlevel_id
            insert_stmt = text("""
                INSERT INTO dim_jobtitles(name, manlevel_id)
                VALUES (:name, :manlevel_id)
                RETURNING id
            """)
            new_id = db_session.execute(insert_stmt, {
                "name": title,
                "manlevel_id": manlevel_id
            }).scalar()
            db_session.commit()
            jobtitle_map[key] = new_id
            result_ids.append(new_id)

    return result_ids

def get_or_create_dim_id_value_pairs(df, db_session, dim_table, source_column, match_column):
    """
    Looks up or inserts values into a dimension table and returns a list of IDs.
    Includes special handling for dim_countries which requires a subregion_id.
    """
    records = db_session.execute(text(f"SELECT id, {match_column} FROM {dim_table}")).fetchall()
    value_map = {row[1].strip().lower(): row[0] for row in records}

    result_ids = []

    for val in df[source_column]:
        if pd.isna(val) or str(val).strip() == "":
            result_ids.append(None)
            continue

        val_clean = str(val).strip().lower()
        dim_id = value_map.get(val_clean)

        if dim_id:
            result_ids.append(dim_id)
        else:
            if dim_table == "dim_countries":
                insert_stmt = text(f"""
                    INSERT INTO {dim_table}({match_column}, subregion_id)
                    VALUES (:val, 999999) RETURNING id
                """)
                new_id = db_session.execute(insert_stmt, {"val": str(val).strip()}).scalar()
            else:
                insert_stmt = text(f"INSERT INTO {dim_table}({match_column}) VALUES (:val) RETURNING id")
                new_id = db_session.execute(insert_stmt, {"val": str(val).strip()}).scalar()

            db_session.commit()
            value_map[val_clean] = new_id
            result_ids.append(new_id)

    return result_ids

def insert_new_companies(df_unique, db_session):
    """
    Inserts new companies into fact_companies and returns a map of index -> new company_id
    """
    company_id_map = {}

    for _, row in df_unique.iterrows():
        insert_stmt = text("""
            INSERT INTO fact_companies (
                name, comp_domain, comp_phone, comp_linkedin,
                address_id, city_id, state_id, postalcode_id, country_id, industry_id,
                annrev, empsize
            ) VALUES (
                :name, :comp_domain, :comp_phone, :comp_linkedin,
                :address_id, :city_id, :state_id, :postalcode_id, :country_id, :industry_id,
                :annrev, :empsize
            ) RETURNING id
        """)

        result = db_session.execute(insert_stmt, {
            "name": row["name"],
            "comp_domain": row["comp_domain"],
            "comp_phone": row["comp_phone"],
            "comp_linkedin": row["comp_linkedin"],
            "address_id": row.get("address_id"),
            "city_id": row.get("city_id"),
            "state_id": row.get("state_id"),
            "postalcode_id": row.get("postalcode_id"),
            "country_id": row.get("country_id"),
            "industry_id": row.get("industry_id"),
            "annrev": row["annrev"],
            "empsize": row["empsize"]
        })

        db_session.commit()
        company_id = result.scalar()
        company_id_map[row["index"]] = company_id

    return company_id_map

def get_or_create_state_ids(df, db_session):
    """
    Resolves state IDs using both state name and country_id.
    Inserts new state records if needed.
    """
    existing = db_session.execute(text("SELECT id, name, country_id FROM dim_states")).fetchall()
    state_map = {(row.name.strip().lower(), row.country_id): row.id for row in existing}

    result_ids = []

    for _, row in df.iterrows():
        state = str(row['compstate']).strip() if pd.notna(row['compstate']) else ""
        country_id = row['country_id']

        if not state or pd.isna(country_id):
            result_ids.append(None)
            continue

        key = (state.lower(), country_id)
        existing_id = state_map.get(key)

        if existing_id:
            result_ids.append(existing_id)
        else:
            insert_stmt = text("""
                INSERT INTO dim_states(name, country_id)
                VALUES (:name, :country_id)
                RETURNING id
            """)
            new_id = db_session.execute(insert_stmt, {
                "name": state,
                "country_id": country_id
            }).scalar()
            db_session.commit()
            state_map[key] = new_id
            result_ids.append(new_id)

    return result_ids

def check_company_existence(df, db):
    query = text("SELECT name, comp_domain FROM fact_companies")
    existing = db.execute(query).fetchall()
    existing_set = set((row.name.strip().lower(), row.comp_domain.strip().lower()) for row in existing)

    def check(row):
        name = str(row.get("companyname", "")).strip().lower()
        domain = str(row.get("comp_domain", "")).strip().lower()
        return "Update" if (name, domain) in existing_set else "New"

    df.insert(df.columns.get_loc("comp_domain") + 1, "company_status", df.apply(check, axis=1))
    return df

def check_contact_existence(df, db):
    query = text("SELECT emplinkedin, empemail FROM fact_contacts")
    existing = db.execute(query).fetchall()
    existing_set = set((str(row.emplinkedin).strip().lower(), str(row.empemail).strip().lower()) for row in existing)

    def check(row):
        linkedin = str(row.get("emplinkedin", "")).strip().lower()
        email = str(row.get("empemail", "")).strip().lower()
        return "Update" if (linkedin, email) in existing_set else "New"

    df.insert(df.columns.get_loc("empemail") + 1, "contact_status", df.apply(check, axis=1))
    return df

def get_existing_values(table_name, db):
    """Fetch all existing values from a dim table"""
    query = text(f"SELECT name FROM {table_name}")
    result = db.execute(query).fetchall()
    return set(row.name.strip().lower() for row in result)

def validate_column(df, column, valid_values):
    """Check if each value in the column exists in the dimension table, and insert status column right after it.
       Special case: if column is 'emailstatus' or 'managementlevel' and value is empty, treat as 'Exists' (valid).
    """
    status_column = f"{column}_status"

    def get_status(val):
        if column in ["emailstatus", "managementlevel"]:
            if pd.isna(val) or str(val).strip() == "":
                return "Exists"  # treat empty as valid
        return "Exists" if pd.notna(val) and str(val).strip().lower() in valid_values else "Not exists"

    status_series = df[column].apply(get_status)

    # Insert status column next to original column
    col_index = df.columns.get_loc(column)
    df.insert(col_index + 1, status_column, status_series)

    return df

# --------------------------- Validation Logic -----------------------
def run_validation(df):
    with next(get_db()) as db:
        for column, dim_table in VALIDATION_MAP.items():
            if column in df.columns:
                valid_values = get_existing_values(dim_table, db)
                df = validate_column(df, column, valid_values)

        if "companyname" in df.columns and "comp_domain" in df.columns:
            df = check_company_existence(df, db)

        if "emplinkedin" in df.columns and "empemail" in df.columns:
            df = check_contact_existence(df, db)

    return df

def export_to_excel_colored(df):
    buffer = BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name.endswith("_status"):
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{get_column_letter(col_idx)}{row}"]
                if cell.value == "Exists":
                    cell.fill = green_fill
                    cell.font = cell.font.copy(color="000000")  # Black font
                elif cell.value == "Not exists":
                    cell.fill = red_fill
                    cell.font = cell.font.copy(color="FFFFFF")  # White font

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

def style_dataframe(df):
    def highlight_status(val):
        if val == "Exists":
            return "background-color: #C6EFCE; color: black"
        elif val == "Not exists":
            return "background-color: #FF0000; color: white"
        else:
            return ""

    # Only show rows with at least one "Not exists" in status columns
    status_cols = [col for col in df.columns if col.endswith("_status")]
    if status_cols:
        df = df[df[status_cols].apply(lambda row: (row == "Not exists").any(), axis=1)]

    return df.style.applymap(highlight_status, subset=status_cols)

def import_validated_data(df_uploaded):
    df_uploaded['index'] = df_uploaded.index + 1

    df_new = df_uploaded[['index', 'name', 'firstname', 'lastname', 'emplinkedin', 'empemail']].copy()

    with next(get_db()) as db:
        df_lookup = df_uploaded[['index', 'managementlevel', 'jobtitle', 'emailstatus']].copy()
        df_lookup['manlevel_id'] = get_or_create_dim_ids(df_lookup, db, "dim_manlevels", "managementlevel", "manlevel_id")
        df_lookup['jobtitle_id'] = get_or_create_jobtitle_ids(df_lookup, db)
        df_lookup['emailstatus_id'] = get_or_create_dim_ids(df_lookup, db, "dim_emailstatuses", "emailstatus", "emailstatus_id")

    df_new = df_new.merge(df_lookup[['index', 'manlevel_id', 'jobtitle_id', 'emailstatus_id']], on='index', how='left')

    df_companies = df_uploaded[['index', 'companyname', 'comp_domain', 'comp_phone', 'comp_linkedin',
                                'annrev', 'empsize', 'address', 'city', 'country', 'compstate', 'postalcode', 'industry']].copy()
    with next(get_db()) as db:
        matched_ids, df_unmatched_companies = get_existing_company_ids(df_companies, db)

    df_new['company_id'] = df_new['index'].map(matched_ids)

    if not df_unmatched_companies.empty:
        df_unique_unmatched = prepare_unique_companies(df_unmatched_companies)

        with next(get_db()) as db:
            df_unique_unmatched['country_id'] = get_or_create_dim_id_value_pairs(df_unique_unmatched, db, 'dim_countries', 'country', 'name')
            df_unique_unmatched['state_id'] = get_or_create_state_ids(df_unique_unmatched, db)
            df_unique_unmatched['address_id'] = get_or_create_dim_id_value_pairs(df_unique_unmatched, db, 'dim_addresses', 'address', 'name')
            df_unique_unmatched['city_id'] = get_or_create_dim_id_value_pairs(df_unique_unmatched, db, 'dim_cities', 'city', 'name')
            df_unique_unmatched['postalcode_id'] = get_or_create_dim_id_value_pairs(df_unique_unmatched, db, 'dim_postalcodes', 'postalcode', 'name')
            df_unique_unmatched['industry_id'] = get_or_create_dim_id_value_pairs(df_unique_unmatched, db, 'dim_industries', 'industry', 'name')

            company_id_map = insert_new_companies(df_unique_unmatched, db)

        df_new['company_id'] = df_new['company_id'].combine_first(df_new['index'].map(company_id_map))

    with next(get_db()) as db:
        contacts_to_insert = df_new[df_new['company_id'].notna()].copy()

        if not contacts_to_insert.empty:
            existing = db.execute(text("""
                SELECT empemail, company_id FROM fact_contacts
                WHERE empemail IS NOT NULL
            """)).fetchall()

            existing_pairs = set((row.empemail.strip().lower(), row.company_id) for row in existing)

            def is_new_contact(row):
                email = str(row['empemail']).strip().lower()
                company_id = row['company_id']
                return (email, company_id) not in existing_pairs

            filtered_contacts = contacts_to_insert[contacts_to_insert.apply(is_new_contact, axis=1)]
            final_contacts = filtered_contacts.drop(columns=['index'], errors='ignore')
            records = final_contacts.to_dict(orient='records')

            if records:
                insert_stmt = text("""
                    INSERT INTO fact_contacts (
                        name, firstname, lastname, emplinkedin, empemail,
                        jobtitle_id, emailstatus_id, company_id
                    ) VALUES (
                        :name, :firstname, :lastname, :emplinkedin, :empemail,
                        :jobtitle_id, :emailstatus_id, :company_id
                    )
                """)
                db.execute(insert_stmt, records)
                db.commit()
                return len(records)
    return 0

def enrich_and_merge_dim(df_base, df_target, source_column, dim_table, dim_column, id_column, extra_insert_values=None):
    """
    Enriches a dimension table and merges the resulting ID into the target DataFrame,
    using a unified key (for companies) to avoid duplicates and ensure accurate matching.

    Parameters:
    - df_base: DataFrame with 'index' and the source column (e.g., domain, LinkedIn, name).
    - df_target: DataFrame into which the resulting ID should be merged.
    - source_column: The raw value column in df_base.
    - dim_table: The name of the target dimension table in the database.
    - dim_column: The column in the DB to match against.
    - id_column: The ID column to be merged in (e.g., company_id).
    - extra_insert_values: Additional values to include during inserts (e.g., {'subregion_id': 999999}).

    Returns:
    - Merged DataFrame with clean ID values and source column removed.
    """
    df_dim = df_base[['index', source_column]].dropna().copy()
    df_dim[id_column] = None

    # Special logic for fact_companies using a check_key
    if dim_table == 'fact_companies':
        df_dim['check_key'] = df_target[['comp_domain', 'comp_linkedin', 'name']].fillna('').agg('|'.join, axis=1).str.lower().str.strip()

        with engine.begin() as conn:
            result = conn.execute(text("SELECT id, comp_domain, comp_linkedin, name FROM fact_companies"))
            df_existing = pd.DataFrame(result.fetchall(), columns=['id', 'comp_domain', 'comp_linkedin', 'name'])

        df_existing['check_key'] = df_existing[['comp_domain', 'comp_linkedin', 'name']].fillna('').agg('|'.join, axis=1).str.lower().str.strip()

        merged_ids = df_dim.merge(df_existing[['id', 'check_key']], on='check_key', how='left')
        df_dim[id_column] = merged_ids['id']

    else:
        # Standard logic for other dimension tables
        if dim_table == 'dim_states' and 'country_id' in df_base.columns:
            df_dim = df_dim.merge(df_base[['index', 'country_id']], on='index', how='left')

        with engine.begin() as conn:
            for idx, row in df_dim.iterrows():
                value = str(row[source_column])
                query_params = {"val": value}
                query = f"SELECT id FROM {dim_table} WHERE {dim_column} = :val"

                if dim_table == 'dim_states':
                    country_id = row.get('country_id', 999999) or 999999
                    query += " AND country_id = :country_id"
                    query_params["country_id"] = country_id

                result = conn.execute(text(query), query_params).fetchone()

                if result:
                    df_dim.at[idx, id_column] = result[0]
                else:
                    insert_data = {dim_column: str(value)} 
                    if dim_table == 'dim_states':
                        insert_data['country_id'] = row.get('country_id', 999999) or 999999
                    if extra_insert_values:
                        insert_data.update(extra_insert_values)

                    columns = ", ".join(insert_data.keys())
                    placeholders = ", ".join(f":{k}" for k in insert_data)
                    insert_query = text(f"INSERT INTO {dim_table} ({columns}) VALUES ({placeholders}) RETURNING id")
                    new_id = conn.execute(insert_query, insert_data).scalar()
                    df_dim.at[idx, id_column] = new_id

    # Merge and clean up
    merged = df_target.merge(df_dim[['index', id_column]], on='index', how='left', suffixes=('', '_new'))

    if f'{id_column}_new' in merged.columns:
        merged[id_column] = merged[f'{id_column}_new']
        merged.drop(columns=[f'{id_column}_new'], inplace=True)

    merged.drop(columns=[source_column, 'industry'], errors='ignore', inplace=True)
    
    return merged

def enrich_and_merge_states(df_base, df_target):
    """
    Handles enrichment and merging of state_id with required country_id for inserts.

    Parameters:
    - df_base: DataFrame with columns 'index', 'name' (state), and 'country_id'.
    - df_target: The target DataFrame to merge state_id into.

    Returns:
    - Updated df_target with state_id merged in.
    """
    df_states = df_base[['index', 'name', 'country_id']].dropna(subset=['name']).drop_duplicates().copy()
    df_states['state_id'] = None

    with engine.begin() as conn:
        for idx, row in df_states.iterrows():
            state_name = row['name']
            country_id = row['country_id'] if pd.notna(row['country_id']) else 999999

            result = conn.execute(
                text("SELECT id FROM dim_states WHERE name = :name AND country_id = :country_id"),
                {"name": state_name, "country_id": country_id}
            ).fetchone()

            if result:
                df_states.at[idx, 'state_id'] = result[0]
            else:
                new_id = conn.execute(
                    text("INSERT INTO dim_states (name, country_id) VALUES (:name, :country_id) RETURNING id"),
                    {"name": state_name, "country_id": country_id}
                ).scalar()
                df_states.at[idx, 'state_id'] = new_id

    merged = df_target.merge(df_states[['index', 'state_id']], on='index', how='left', suffixes=('', '_new'))

    # Clean up any _new suffix if it appears
    if 'state_id_new' in merged.columns:
        merged['state_id'] = merged['state_id_new']
        merged.drop(columns=['state_id_new'], inplace=True)

    return merged

def extract_lower_bound(size_str):
    """
    Extract the lower numeric bound from a size string (e.g., '51-100' → 51).
    Returns None if no match found.
    """
    if pd.isna(size_str):
        return None
    size_str = size_str.replace(',', '').lower()
    match = re.search(r'\d+', size_str)
    if match:
        return int(match.group())
    return None

def extract_revenue_lower_bound(revenue_str):
    """
    Extracts the lower bound of company revenue and converts it into an integer.
    Supports formats like '$10M', '5B', etc. Defaults to millions if unit is missing.
    """
    if pd.isna(revenue_str):
        return None
    revenue_str = revenue_str.replace(',', '').replace('$', '').strip().upper()

    match = re.match(r'(\d+)([MB])', revenue_str)
    if match:
        number = int(match.group(1))
        unit = match.group(2)
    else:
        match = re.search(r'(\d+)', revenue_str)
        if not match:
            return None
        number = int(match.group(1))
        unit = 'M'  # Assume millions if no unit

    multiplier = 1_000_000 if unit == 'M' else 1_000_000_000
    return number * multiplier

def match_companies_by_domain_or_linkedin(df_companies):
    df_companies = df_companies.copy()

    # Normalize individual fields
    df_companies['normalized_domain'] = df_companies['comp_domain'].fillna('').apply(normalize_domain)
    df_companies['normalized_linkedin'] = df_companies['comp_linkedin'].fillna('').apply(normalize_domain)
    df_companies['normalized_name'] = df_companies['name'].fillna('').str.lower().str.strip()

    # Build match key
    df_companies['match_key'] = df_companies[['normalized_domain', 'normalized_linkedin', 'normalized_name']].agg('|'.join, axis=1)

    with engine.begin() as conn:
        result = conn.execute(text("SELECT id AS company_id, comp_domain, comp_linkedin, name FROM fact_companies"))
        df_existing = pd.DataFrame(result.fetchall(), columns=['company_id', 'comp_domain', 'comp_linkedin', 'name'])

    # Normalize database columns before building match_key
    df_existing['normalized_domain'] = df_existing['comp_domain'].fillna('').apply(normalize_domain)
    df_existing['normalized_linkedin'] = df_existing['comp_linkedin'].fillna('').apply(normalize_domain)
    df_existing['normalized_name'] = df_existing['name'].fillna('').str.lower().str.strip()

    # Now safely build the match key
    df_existing['match_key'] = df_existing[['normalized_domain', 'normalized_linkedin', 'normalized_name']].agg('|'.join, axis=1)

    # Perform the merge
    df_matched = df_companies.merge(df_existing[['company_id', 'match_key']], on='match_key', how='left')

    # Clean up
    df_matched.drop(columns=['match_key', 'normalized_domain', 'normalized_linkedin', 'normalized_name'], inplace=True)

    return df_matched

def insert_unmatched_companies(df_companies):
    """
    Inserts new companies from df_companies where company_id is missing.
    Retrieves new IDs and merges them back into the DataFrame.
    Assumes df_companies has all required DB fields and 'index'.
    """
    df = df_companies.copy()

    # Step 1: Filter unmatched companies
    df_new = df[df['company_id'].isna()].copy()
    if df_new.empty:
        return df  # Nothing to insert

    # Step 2: Drop columns not in DB
    df_to_insert = df_new.drop(columns=['index', 'company_id'])

    inserted_ids = []

    with engine.begin() as conn:
        insert_stmt = text("""
            INSERT INTO fact_companies (
                name, comp_domain, empsize, annrev, comp_linkedin, comp_phone,
                industry_id, country_id, state_id, city_id, postalcode_id, address_id
            ) VALUES (
                :name, :comp_domain, :empsize, :annrev, :comp_linkedin, :comp_phone,
                :industry_id, :country_id, :state_id, :city_id, :postalcode_id, :address_id
            ) RETURNING id
        """)

        for _, row in df_to_insert.iterrows():
            row_dict = row.to_dict()
            index_val = df_new.at[row.name, 'index']
            new_id = conn.execute(insert_stmt, row_dict).scalar()
            inserted_ids.append({'index': index_val, 'company_id': new_id})

    # Step 3: Merge new IDs back into df_companies
    inserted_df = pd.DataFrame(inserted_ids)
    df = df.merge(inserted_df, on='index', how='left', suffixes=('', '_new'))

    # Fill missing company_id from inserted
    df['company_id'] = df['company_id'].fillna(df['company_id_new'])
    df.drop(columns=['company_id_new'], inplace=True)

    return df

def normalize_domain(domain):
    """
    Cleans and normalizes domain names by removing protocols and www prefixes.
    """
    if pd.isna(domain):
        return None
    domain = domain.lower().strip()
    domain = re.sub(r'^https?://', '', domain)
    domain = re.sub(r'^www\.', '', domain)
    domain = domain.strip('/')
    return domain

def process_campaign_updates(file):
    import pandas as pd
    import re
    from sqlalchemy import text
    from app_backend.database import engine

    def normalize_linkedin(url):
        if pd.isna(url):
            return ""
        url = str(url).strip().lower()
        url = re.sub(r'^https?://(www\.)?', '', url)
        return url

    if file.name.endswith(".csv"):
        df_base = pd.read_csv(file)
    elif file.name.endswith((".xls", ".xlsx")):
        df_base = pd.read_excel(file)
    else:
        raise ValueError("Unsupported file format.")

    df_base['index'] = df_base.index + 1
    total_records = len(df_base)

    df_base = df_base[[ 
        'index', 'company_name', 'domain', 'company_size',
        'company_revenue', 'industry', 'company_linkedin_link', 'first_name', 'last_name',
        'title', 'email', 'contact_linkedin_link', 'work_phone_number', 'street',
        'city', 'state', 'country', 'zip_code', 'qa_disposition'
    ]]

    df_companies = df_base[[ 
        'index', 'company_name', 'domain', 'company_size', 'company_revenue',
        'industry', 'company_linkedin_link', 'work_phone_number'
    ]].copy()

    for col in ['address_id', 'city_id', 'state_id', 'postalcode_id', 'country_id', 'industry_id']:
        df_companies[col] = None

    df_addresses = df_base[['index', 'street']].dropna().rename(columns={'street': 'name'})
    df_cities = df_base[['index', 'city']].dropna().rename(columns={'city': 'name'})
    df_postalcodes = df_base[['index', 'zip_code']].dropna().rename(columns={'zip_code': 'name'})
    df_countries = df_base[['index', 'country']].dropna().rename(columns={'country': 'name'})
    df_industries = df_base[['index', 'industry']].dropna().rename(columns={'industry': 'name'})
    df_states = df_base[['index', 'state', 'country']].dropna(subset=['state']).rename(columns={'state': 'name', 'country': 'country_name'})

    df_companies = enrich_and_merge_dim(df_addresses, df_companies, 'name', 'dim_addresses', 'name', 'address_id')
    df_companies = enrich_and_merge_dim(df_cities, df_companies, 'name', 'dim_cities', 'name', 'city_id')
    df_companies = enrich_and_merge_dim(df_postalcodes, df_companies, 'name', 'dim_postalcodes', 'name', 'postalcode_id')
    df_companies = enrich_and_merge_dim(df_industries, df_companies, 'name', 'dim_industries', 'name', 'industry_id')
    df_companies = enrich_and_merge_dim(df_countries, df_companies, 'name', 'dim_countries', 'name', 'country_id', extra_insert_values={'subregion_id': 999999})

    df_states = df_states.merge(df_companies[['index', 'country_id']], on='index', how='left')
    df_companies = enrich_and_merge_states(df_states, df_companies)

    df_companies = df_companies.rename(columns={
        'company_name': 'name',
        'domain': 'comp_domain',
        'company_size': 'empsize',
        'company_revenue': 'annrev',
        'company_linkedin_link': 'comp_linkedin',
        'work_phone_number': 'comp_phone'
    })

    df_companies['empsize'] = df_companies['empsize'].apply(lambda x: x if pd.api.types.is_integer(x) else extract_lower_bound(x))
    df_companies['annrev'] = df_companies['annrev'].apply(lambda x: x if pd.api.types.is_integer(x) else extract_revenue_lower_bound(x))

    df_companies = match_companies_by_domain_or_linkedin(df_companies)
    df_companies = update_matched_companies_if_different(
        df_companies[df_companies['company_id'].notna()].copy(), df_companies
    )
    df_companies = insert_unmatched_companies(df_companies)
    df_companies = df_companies[df_companies['company_id'].notna()].copy()
    df_companies = df_companies.drop_duplicates(subset='index', keep='last')

    df_contacts = df_base[[ 
        'index', 'first_name', 'last_name', 'title', 'email',
        'contact_linkedin_link', 'work_phone_number', 'qa_disposition'
    ]].copy()

    df_contacts['name'] = df_contacts['first_name'].fillna('') + ' ' + df_contacts['last_name'].fillna('')
    df_contacts = df_contacts.merge(df_companies[['index', 'company_id']], on='index', how='left')

    df_contacts = df_contacts.rename(columns={
        'first_name': 'firstname',
        'last_name': 'lastname',
        'title': 'jobtitle',
        'email': 'empemail',
        'contact_linkedin_link': 'emplinkedin',
        'work_phone_number': 'empphone'
    })

    df_contacts['empphone'] = df_contacts['empphone'].fillna('').apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x).strip())
    df_contacts['qa_disposition'] = df_contacts['qa_disposition'].fillna('').str.strip().str.title()
    df_contacts['emailstatus_id'] = df_contacts['qa_disposition'].apply(lambda x: 1 if x == 'Qualified' else 4)
    df_contacts.drop(columns=['qa_disposition'], inplace=True)
    df_contacts[['address_id', 'city_id', 'postalcode_id', 'country_id', 'state_id']] = 999999

    df_contacts = df_contacts[~((df_contacts['empemail'].isna() | (df_contacts['empemail'].astype(str).str.strip() == "")) &
                                (df_contacts['emplinkedin'].isna() | (df_contacts['emplinkedin'].astype(str).str.strip() == "")))].copy()

    df_jobtitles = df_contacts[['index', 'jobtitle']].dropna().copy()
    df_jobtitles['jobtitle_id'] = None

    with engine.begin() as conn:
        for idx, row in df_jobtitles.iterrows():
            title = row['jobtitle']
            result = conn.execute(text("SELECT id FROM dim_jobtitles WHERE name = :name"), {"name": title}).fetchone()
            if result:
                df_jobtitles.at[idx, 'jobtitle_id'] = result[0]
            else:
                new_id = conn.execute(text("INSERT INTO dim_jobtitles (name) VALUES (:name) RETURNING id"), {"name": title}).scalar()
                df_jobtitles.at[idx, 'jobtitle_id'] = new_id

    df_contacts = df_contacts.merge(df_jobtitles[['index', 'jobtitle_id']], on='index', how='left')
    df_contacts.drop(columns=['jobtitle'], inplace=True)

    df_contacts['emplinkedin'] = df_contacts['emplinkedin'].fillna('').apply(normalize_linkedin)
    df_contacts['empemail'] = df_contacts['empemail'].fillna('').astype(str).str.strip().str.lower()

    with engine.begin() as conn:
        result = conn.execute(text("SELECT id AS contact_id, empemail, emplinkedin, name FROM fact_contacts"))
        df_existing_contacts = pd.DataFrame(result.fetchall(), columns=['contact_id', 'empemail', 'emplinkedin', 'name'])

    df_existing_contacts['emplinkedin'] = df_existing_contacts['emplinkedin'].fillna('').apply(normalize_linkedin)
    df_existing_contacts['empemail'] = df_existing_contacts['empemail'].fillna('').astype(str).str.strip().str.lower()

    df_existing_contacts['composite_key'] = df_existing_contacts.apply(
        lambda row: f"{str(row['empemail']).strip().lower()}|{normalize_linkedin(row['emplinkedin']) if pd.notna(row['emplinkedin']) else ''}",
        axis=1
    )

    email_map = df_existing_contacts.set_index('empemail')['contact_id'].to_dict()
    linkedin_map = df_existing_contacts.set_index('emplinkedin')['contact_id'].to_dict()
    composite_map = df_existing_contacts.set_index('composite_key')['contact_id'].to_dict()

    df_contacts['contact_id'] = None
    for idx, row in df_contacts.iterrows():
        email = str(row['empemail']).strip().lower() if pd.notna(row['empemail']) else ''
        linkedin = normalize_linkedin(row['emplinkedin']) if pd.notna(row['emplinkedin']) else ''
        key = f"{email}|{linkedin}"
        contact_id = composite_map.get(key)
        if not contact_id and email:
            contact_id = email_map.get(email)
        if not contact_id and linkedin:
            contact_id = linkedin_map.get(linkedin)
        df_contacts.at[idx, 'contact_id'] = contact_id

    # ✅ MOVED BELOW emailstatus_id mapping
    df_matched_contacts = df_contacts[df_contacts['contact_id'].notna()].copy()
    df_matched_contacts, updated_records = update_matched_contacts_if_different(df_matched_contacts)

    df_new_contacts = df_contacts[df_contacts['contact_id'].isna()].copy()
    if not df_new_contacts.empty:
        df_to_insert = df_new_contacts.drop(columns=['index', 'contact_id'])
        inserted_ids = []
        with engine.begin() as conn:
            insert_stmt = text("""
                INSERT INTO fact_contacts (
                    name, firstname, lastname, empemail, emplinkedin, empphone, emailstatus_id,
                    address_id, city_id, postalcode_id, country_id, state_id,
                    company_id, jobtitle_id
                ) VALUES (
                    :name, :firstname, :lastname, :empemail, :emplinkedin, :empphone, :emailstatus_id,
                    :address_id, :city_id, :postalcode_id, :country_id, :state_id,
                    :company_id, :jobtitle_id
                ) RETURNING id
            """)
            for _, row in df_to_insert.iterrows():
                row_dict = row.to_dict()
                new_id = conn.execute(insert_stmt, row_dict).scalar()
                inserted_ids.append({'index': row.name, 'contact_id': new_id})

        inserted_df = pd.DataFrame(inserted_ids)
        df_contacts = df_contacts.merge(inserted_df, left_index=True, right_on='index', how='left', suffixes=('', '_new'))
        df_contacts['contact_id'] = df_contacts['contact_id'].fillna(df_contacts['contact_id_new'])
        df_contacts.drop(columns=['contact_id_new', 'index_y'], inplace=True, errors='ignore')
        df_contacts.rename(columns={'index_x': 'index'}, inplace=True)
        inserted_records = len(df_to_insert)
    else:
        inserted_records = 0

    df_contacts = pd.concat([df_matched_contacts, df_contacts[df_contacts['contact_id'].notna()]], ignore_index=True)
    df_contacts = df_contacts.drop_duplicates(subset='index', keep='first')

    return {
        "total": total_records,
        "inserted": inserted_records,
        "updated": updated_records
    }

def update_matched_companies_if_different(df_matched, df_all_columns):
    """
    For matched companies: compare each relevant column to the DB.
    If there's at least one difference, issue an UPDATE.

    Parameters:
    - df_matched: DataFrame with company_id already matched
    - df_all_columns: full df_companies with all columns (to compare)

    Returns:
    - DataFrame of matched companies (with updates applied if needed)
    """
    columns_to_check = [
        'name', 'comp_domain', 'comp_phone', 'comp_linkedin',
        'address_id', 'city_id', 'state_id', 'postalcode_id', 'country_id', 'industry_id',
        'annrev', 'empsize'
    ]

    # Get unique company_ids from matched
    matched = df_matched[df_matched['company_id'].notna()].copy()

    # Load current values from DB for these company_ids
    company_ids = matched['company_id'].dropna().unique().tolist()
    query = text(f"""
        SELECT id AS company_id, name, comp_domain, comp_phone, comp_linkedin,
               address_id, city_id, state_id, postalcode_id, country_id, industry_id,
               annrev, empsize
        FROM fact_companies
        WHERE id = ANY(:ids)
    """)
    with engine.begin() as conn:
        db_data = pd.DataFrame(conn.execute(query, {"ids": company_ids}).fetchall(), columns=['company_id'] + columns_to_check)

    # Merge matched with db_data to compare
    compare_df = matched.merge(db_data, on='company_id', suffixes=('', '_db'))

    updates = []
    for _, row in compare_df.iterrows():
        differences = {}
        for col in columns_to_check:
            if col in ['comp_domain', 'comp_linkedin']:
                val_df = normalize_domain(row[col])
                val_db = normalize_domain(row[f"{col}_db"])
            else:
                val_df = normalize_value(row[col])
                val_db = normalize_value(row[f"{col}_db"])
            if val_df != val_db:
                differences[col] = row[col]  # use original value for update
        if differences:
            differences['company_id'] = row['company_id']
            updates.append(differences)

    # Perform updates if needed
    if updates:
        with engine.begin() as conn:
            for update_data in updates:
                company_id = update_data.pop('company_id')
                set_clause = ", ".join([f"{k} = :{k}" for k in update_data])
                update_stmt = text(f"""
                    UPDATE fact_companies SET {set_clause} WHERE id = :company_id
                """)
                update_data['company_id'] = company_id
                conn.execute(update_stmt, update_data)

    return matched

def normalize_value(val):
    if pd.isna(val) or val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    try:
        return str(val).strip().lower()
    except:
        return str(val)

def update_matched_contacts_if_different(df_matched_contacts):
    """
    Checks for differences in matched contacts and updates records in fact_contacts
    if any field is different.
    """
    # Filter out any records with missing company_id
    df_matched_contacts = df_matched_contacts[df_matched_contacts['company_id'].notna()].copy()

    columns_to_check = [
        'name', 'firstname', 'lastname', 'empemail', 'emplinkedin', 'empphone', 'emailstatus_id',
        'address_id', 'city_id', 'postalcode_id', 'country_id', 'state_id',
        'company_id', 'jobtitle_id'
    ]

    contact_ids = df_matched_contacts['contact_id'].dropna().unique().tolist()

    query = text(f"""
        SELECT id AS contact_id, {', '.join(columns_to_check)}
        FROM fact_contacts
        WHERE id = ANY(:ids)
    """)
    with engine.begin() as conn:
        db_data = pd.DataFrame(conn.execute(query, {"ids": contact_ids}).fetchall(), columns=['contact_id'] + columns_to_check)

    compare_df = df_matched_contacts.merge(db_data, on='contact_id', suffixes=('', '_db'))

    updates = []
    for _, row in compare_df.iterrows():
        differences = {}
        # ✅ Add this sanity check
        for col in columns_to_check:
            val_df_raw = row[col]
            val_db_raw = row[f"{col}_db"]

            val_df = normalize_value(val_df_raw)
            val_db = normalize_value(val_db_raw)

            if val_df != val_db:
                differences[col] = row[col]

        if differences:
            differences['contact_id'] = row['contact_id']
            updates.append(differences)

    if updates:
        with engine.begin() as conn:
            for update_data in updates:
                contact_id = update_data.pop('contact_id')
                set_clause = ", ".join([f"{k} = :{k}" for k in update_data])
                update_stmt = text(f"""
                    UPDATE fact_contacts SET {set_clause} WHERE id = :contact_id
                """)
                update_data['contact_id'] = contact_id
                conn.execute(update_stmt, update_data)

    return df_matched_contacts, len(updates)

def update_cached_contacts(changed_ids: list[int]) -> bool:
    """
    Incrementally refresh cached_full_contacts_data AND cached_filters_contacts_data from the view.
    For each changed id:
    - Insert if id is not in cache
    - Update if id already exists
    """
    from sqlalchemy import text
    import pandas as pd
    from datetime import datetime
    from app_backend.database import engine

    print(f"📥 Refreshing cache for {len(changed_ids)} contact(s): {changed_ids}")

    if not changed_ids:
        print("⚠️ No changed IDs provided.")
        return True

    try:
        id_list = ', '.join(str(i) for i in changed_ids)
        query = f"SELECT * FROM Vw_full_contacts_data WHERE id IN ({id_list})"

        print(f"📄 Running query: {query}")
        df = pd.read_sql(query, engine)
        print(f"📊 View query returned {len(df)} rows.")
        print(df[['id']].head())
        df['last_updated'] = datetime.now()

        def upsert_into_table(table_name: str):
            with engine.begin() as conn:
                for _, row in df.iterrows():
                    columns = df.columns.tolist()
                    placeholders = ', '.join(f":{col}" for col in columns)
                    update_stmt = ', '.join(f"{col} = EXCLUDED.{col}" for col in columns if col != 'id')

                    upsert_query = text(f"""
                        INSERT INTO {table_name} ({', '.join(columns)})
                        VALUES ({placeholders})
                        ON CONFLICT (id) DO UPDATE
                        SET {update_stmt}
                    """)
                    conn.execute(upsert_query, row.to_dict())

        upsert_into_table("cached_full_contacts_data")
        upsert_into_table("cached_filters_contacts_data")

        print(f"✅ {len(df)} record(s) upserted into both cache tables.")
        return True

    except Exception as e:
        print(f"❌ Failed to update cache: {e}")
        return False

def trigger_background_refresh():
    try:
        venv_python = "/home/ubuntu/aws-ff-data/aws-ff-data-env/bin/python"  # Adjust as needed
        result = subprocess.run(
            [venv_python, "-m", "app_backend.refresh_cached"],
            capture_output=True,
            text=True,
            timeout=120
        )
        if result.returncode == 0:
            st.success("✅ Cache refresh triggered.")
        else:
            st.warning(f"⚠️ Cache refresh ran but returned error:\n{result.stderr}")
    except Exception as e:
        st.error(f"❌ Failed to run background refresh: {e}")
        
def safe_exit(message):
    if 'ipykernel' in sys.modules:  # running in Jupyter
        raise RuntimeError(message)
    else:
        print(message)
        sys.exit()

def replace_nan_with_empty_string(df: pd.DataFrame, columns: list) -> pd.DataFrame:
    """
    Replace NaN values with empty strings in the specified columns of a DataFrame.

    Parameters:
    - df: pandas DataFrame
    - columns: list of column names to process

    Returns:
    - Updated DataFrame with NaNs replaced by empty strings in specified columns
    """
    for col in columns:
        if col in df.columns:
            df[col] = df[col].fillna('')
    return df

def convert_qa_disposition(df: pd.DataFrame, column: str = 'qa_disposition') -> pd.DataFrame:
    """
    Convert qa_disposition to numeric values:
    - 'Qualified' => 1
    - anything else => 4

    Parameters:
    - df: pandas DataFrame
    - column: name of the disposition column (default is 'qa_disposition')

    Returns:
    - Updated DataFrame with numeric qa_disposition
    """
    if column in df.columns:
        df[column] = df[column].apply(lambda x: 1 if str(x).strip().lower() == 'qualified' else 4)
    return df

def convert_zipcode_to_string(df: pd.DataFrame, column: str = 'comp_zipcode') -> pd.DataFrame:
    """
    Ensures the zipcode column is converted to string, preserving leading zeros and mixed formats.

    Parameters:
    - df: pandas DataFrame
    - column: name of the zipcode column

    Returns:
    - Updated DataFrame with zipcode as string
    """
    if column in df.columns:
        df[column] = df[column].astype(str).fillna('').str.strip()
    return df

def truncate_linkedin_fields_with_log(df: pd.DataFrame, columns: list, max_length: int = 255) -> pd.DataFrame:
    """
    Truncate values in specified columns to max_length characters and print count of affected rows.

    Parameters:
    - df: pandas DataFrame
    - columns: list of column names to truncate
    - max_length: maximum length allowed (default: 255)

    Returns:
    - Updated DataFrame with truncated values
    """
    for col in columns:
        if col in df.columns:
            original_lengths = df[col].astype(str).str.len()
            affected_count = (original_lengths > max_length).sum()
            if affected_count > 0:
                print(f"🔧 Truncated {affected_count} rows in column '{col}' to {max_length} characters.")
                df[col] = df[col].astype(str).str.slice(0, max_length)
    return df


import pandas as pd
import re

def clean_urls(df: pd.DataFrame, columns: list) -> pd.DataFrame:
    """
    Clean URL-like fields by:
    - Removing 'http://' or 'https://'
    - Removing 'www.', 'www1.', 'www2.', etc.
    - Removing trailing slashes

    Parameters:
    - df: pandas DataFrame
    - columns: list of column names to clean

    Returns:
    - Updated DataFrame with cleaned URL fields
    """
    http_pattern = re.compile(r'^https?://', re.IGNORECASE)
    www_pattern = re.compile(r'^www[0-9]*\.', re.IGNORECASE)

    for col in columns:
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .apply(lambda x: www_pattern.sub('', http_pattern.sub('', x.strip())) if x else '')
                .str.rstrip('/')
            )
    return df


import pandas as pd
from sqlalchemy import text

def enrich_and_merge_dim(df: pd.DataFrame, column: str, dim_table: str, db_session) -> pd.DataFrame:
    """
    Maps values in df[column] to IDs from dim_table.
    - Assumes df[column] is already cleaned
    - Inserts missing values into dim_table
    - Returns df with an added column: {column}_id
    """
    from sqlalchemy import text

    # 1. Unique values
    unique_vals = df[column].drop_duplicates().tolist()

    # 2. Query existing IDs
    placeholders = ', '.join([f":val{i}" for i in range(len(unique_vals))])
    values = {f"val{i}": v for i, v in enumerate(unique_vals)}

    query = text(f"SELECT id, name FROM {dim_table} WHERE name IN ({placeholders})")
    result = db_session.execute(query, values).fetchall()
    id_map = {row.name: row.id for row in result}

    # 3. Insert missing values
    missing_vals = [v for v in unique_vals if v not in id_map]
    for val in missing_vals:
        if dim_table == "dim_countries":
            new_id = db_session.execute(
                text(f"INSERT INTO {dim_table} (name, subregion_id) VALUES (:name, 999999) RETURNING id"),
                {"name": val}
            ).scalar()
        else:
            new_id = db_session.execute(
                text(f"INSERT INTO {dim_table} (name) VALUES (:name) RETURNING id"),
                {"name": val}
            ).scalar()
        id_map[val] = new_id
    db_session.commit()

    # 4. Map all values back to df
    df[f"{column}_id"] = df[column].map(id_map)
    return df

def enrich_and_merge_dim_old(df: pd.DataFrame, column: str, dim_table: str, db_session) -> pd.DataFrame:
    from sqlalchemy import text

    # Extract distinct values and clean them
    df_sub = df[[column]].copy()
    df_sub[column] = df_sub[column].astype(str).fillna('Unknown').str.strip()
    df_sub = df_sub.reset_index()

    unique_values = df_sub[column].drop_duplicates().tolist()

    # Step 1: Query existing values
    placeholders = ', '.join([f":val{i}" for i in range(len(unique_values))])
    values = {f"val{i}": v for i, v in enumerate(unique_values)}

    existing_query = text(f"SELECT id, name FROM {dim_table} WHERE name IN ({placeholders})")
    existing_rows = db_session.execute(existing_query, values).fetchall()
    df_existing = pd.DataFrame(existing_rows, columns=['id', 'name'])

    # Step 2: Find missing and insert
    existing_names = set(df_existing['name'].tolist())
    missing = [v for v in unique_values if v not in existing_names]

    new_rows = []
    for name in missing:
        insert_stmt = text(f"INSERT INTO {dim_table} (name) VALUES (:name) RETURNING id")
        new_id = db_session.execute(insert_stmt, {"name": name}).scalar()
        new_rows.append((new_id, name))
    db_session.commit()

    df_new = pd.DataFrame(new_rows, columns=['id', 'name'])
    df_all = pd.concat([df_existing, df_new], ignore_index=True)

    # Step 3: Merge back to original
    df_sub = df_sub.merge(df_all, left_on=column, right_on='name', how='left')
    df_sub = df_sub.set_index('index')
    df_sub = df_sub.rename(columns={'id': f"{column}_id"})

    return df_sub[[column, f"{column}_id"]]


def enrich_and_merge_states(df: pd.DataFrame, state_col: str, country_col: str, db_session) -> pd.DataFrame:
    """
    Enrich states using both state and country.
    Insert missing state-country pairs into dim_states.
    Requires df to already have {country_col}_id populated.

    Parameters:
    - df: main DataFrame
    - state_col: e.g., 'comp_state'
    - country_col: e.g., 'comp_country'
    - db_session: SQLAlchemy session

    Returns:
    - DataFrame with original state_col and {state_col}_id
    """
    from sqlalchemy import text

    id_col = f'{state_col}_id'
    country_id_col = f'{country_col}_id'
    df_sub = df[[state_col, country_id_col]].copy().reset_index()

    # Handle empty state values
    df_sub[state_col] = df_sub[state_col].apply(lambda x: x if str(x).strip() else '__EMPTY__')

    df_unique = df_sub.drop_duplicates()

    # Query existing states
    query = text(f"""
        SELECT id, name, country_id FROM dim_states
        WHERE (name, country_id) IN (
            SELECT UNNEST(:names) AS name, UNNEST(:cids) AS country_id
        )
    """)
    names = df_unique[state_col].tolist()
    cids = df_unique[country_id_col].tolist()
    result = db_session.execute(query, {'names': names, 'cids': cids}).fetchall()
    df_existing = pd.DataFrame(result, columns=['id', 'name', 'country_id'])

    # Identify missing
    df_merged = df_unique.merge(df_existing, left_on=[state_col, country_id_col], right_on=['name', 'country_id'], how='left')
    missing = df_merged[df_merged['id'].isna()][[state_col, country_id_col]].drop_duplicates()

    # Insert missing (handle empty)
    new_ids = []
    for _, row in missing.iterrows():
        name = row[state_col]
        cid = row[country_id_col]
        if name == '__EMPTY__':
            new_ids.append((999999, '__EMPTY__', cid))
        else:
            insert_stmt = text(f"""
                INSERT INTO dim_states (name, country_id)
                VALUES (:name, :country_id)
                RETURNING id
            """)
            result = db_session.execute(insert_stmt, {'name': name, 'country_id': cid})
            new_id = result.fetchone()[0]
            new_ids.append((new_id, name, cid))
    db_session.commit()

    # Merge all back
    df_new = pd.DataFrame(new_ids, columns=['id', 'name', 'country_id'])
    df_all = pd.concat([df_existing, df_new], ignore_index=True)

    df_sub = df_sub.merge(df_all, left_on=[state_col, country_id_col], right_on=['name', 'country_id'], how='left')
    df_sub = df_sub.drop(columns=['name', 'country_id']).set_index('index')
    df_sub[state_col] = df_sub[state_col].replace('__EMPTY__', '')
    df_sub = df_sub.rename(columns={'id': id_col})
    return df_sub[[state_col, id_col]], len(df_unique), len(df_new)

def get_company_ids(df_companies: pd.DataFrame, db_session) -> pd.DataFrame:
    """
    Match companies in df_companies with fact_companies by (name, domain, linkedin),
    using NULL-safe comparison. Returns df_companies with company_id column.
    """
    from sqlalchemy import text

    # Deduplicate keys for lookup
    unique_keys = df_companies[['name', 'comp_domain', 'comp_linkedin']].drop_duplicates()

    results = []

    for _, row in unique_keys.iterrows():
        query = text("""
            SELECT id
            FROM fact_companies
            WHERE name IS NOT DISTINCT FROM :name
              AND comp_domain IS NOT DISTINCT FROM :comp_domain
              AND comp_linkedin IS NOT DISTINCT FROM :comp_linkedin
            LIMIT 1
        """)
        result = db_session.execute(query, {
            'name': row['name'],
            'comp_domain': row['comp_domain'],
            'comp_linkedin': row['comp_linkedin']
        }).fetchone()

        if result:
            results.append({
                'name': row['name'],
                'comp_domain': row['comp_domain'],
                'comp_linkedin': row['comp_linkedin'],
                'company_id': result[0]
            })

    df_matches = pd.DataFrame(results)

    # Merge back to original
    df_companies = df_companies.merge(
        df_matches,
        on=['name', 'comp_domain', 'comp_linkedin'],
        how='left'
    )

    return df_companies

def compare_companies_to_db(df_companies: pd.DataFrame, db_session) -> pd.DataFrame:
    """
    Compare each company row with fact_companies, normalize values,
    and label each row as 'Insert', 'Update', or 'No Update'.

    Adds 'id' column (same as company_id) and includes it in comparison.
    Ensures comp_phone is treated as string for comparison.
    """
    from sqlalchemy import text

    df_companies['id'] = df_companies['company_id']  # duplicate column

    status_list = []

    compare_fields = [
        'id', 'name', 'comp_domain', 'comp_linkedin', 'comp_phone',
        'annrev', 'empsize', 'address_id', 'country_id',
        'postalcode_id', 'city_id', 'state_id', 'industry_id'
    ]

    for _, row in df_companies.iterrows():
        company_id = row.get('company_id')

        if pd.isna(company_id):
            status_list.append('Insert')
            continue

        query = text(f"""
            SELECT id, name, comp_domain, comp_linkedin, comp_phone,
                   annrev, empsize, address_id, country_id,
                   postalcode_id, city_id, state_id, industry_id
            FROM fact_companies
            WHERE id = :company_id
        """)
        result = db_session.execute(query, {'company_id': int(company_id)}).fetchone()

        if not result:
            status_list.append('Insert')  # orphaned ID
            continue

        db_row = dict(zip(compare_fields, result))

        is_different = False
        for field in compare_fields:
            df_val = row.get(field)
            db_val = db_row.get(field)

            # Normalize string fields
            if field == 'comp_phone':
                df_val = str(df_val).strip() if pd.notna(df_val) else None
                db_val = str(db_val).strip() if pd.notna(db_val) else None
                if df_val == '':
                    df_val = None
                if db_val == '':
                    db_val = None
            else:
                # Treat empty string as None
                if isinstance(df_val, str) and df_val.strip() == '':
                    df_val = None
                if isinstance(db_val, str) and db_val.strip() == '':
                    db_val = None

                # Treat NaN as None
                if pd.isna(df_val):
                    df_val = None
                if pd.isna(db_val):
                    db_val = None

            if df_val != db_val:
                is_different = True
                break

        status_list.append('Update' if is_different else 'No Update')

    df_companies['status'] = status_list
    return df_companies

def upsert_companies(df_companies: pd.DataFrame, db_session) -> pd.DataFrame:
    """
    Updates existing companies (status = 'Update') and inserts new ones (status = 'Insert').
    For inserts, retrieves and assigns new company_id values using the row index.

    Returns:
        Updated df_companies with new company_id/id values after insert.
    """
    from sqlalchemy import text

    update_fields = [
        'name', 'comp_domain', 'comp_linkedin', 'comp_phone',
        'annrev', 'empsize', 'address_id', 'country_id',
        'postalcode_id', 'city_id', 'state_id', 'industry_id'
    ]

    df_result = df_companies.copy()

    # 1. Updates
    df_update = df_result[df_result['status'] == 'Update']
    for _, row in df_update.iterrows():
        set_clause = ', '.join([f"{field} = :{field}" for field in update_fields])
        query = text(f"""
            UPDATE fact_companies
            SET {set_clause}
            WHERE id = :id
        """)
        params = {field: row[field] for field in update_fields}
        params['id'] = int(row['id'])
        db_session.execute(query, params)

    # 2. Inserts (match back by DataFrame index)
    df_insert = df_result[df_result['status'] == 'Insert'].copy()
    new_ids = []

    if not df_insert.empty:
        for i, row in df_insert.iterrows():
            insert_query = text(f"""
                INSERT INTO fact_companies (
                    {', '.join(update_fields)}
                ) VALUES (
                    {', '.join([f":{field}" for field in update_fields])}
                ) RETURNING id
            """)
            params = {field: row[field] if pd.notna(row[field]) else None for field in update_fields}
            result = db_session.execute(insert_query, params)
            new_id = result.scalar()
            new_ids.append((i, new_id))

    db_session.commit()

    # Assign new IDs using the original DataFrame index
    for idx, company_id in new_ids:
        df_result.at[idx, 'company_id'] = company_id
        df_result.at[idx, 'id'] = company_id

    print(f"✅ {len(df_update)} rows updated, {len(new_ids)} rows inserted into fact_companies.")
    return df_result

def get_contact_ids(df_contacts: pd.DataFrame, db_session) -> pd.DataFrame:
    """
    Match contacts to fact_contacts using empemail + emplinkedin.
    Returns df_contacts with contact_id column.
    If no matches found at all, returns contact_id = None for all.
    """
    from sqlalchemy import text

    df_contacts = df_contacts.copy()

    # Normalize input
    for col in ['empemail', 'emplinkedin']:
        df_contacts[col] = df_contacts[col].replace('', None)
        df_contacts[col] = df_contacts[col].where(pd.notna(df_contacts[col]), None)

    keys_df = df_contacts[['index', 'empemail', 'emplinkedin']].drop_duplicates()
    results = []

    for _, row in keys_df.iterrows():
        result = db_session.execute(text("""
            SELECT id FROM fact_contacts
            WHERE empemail IS NOT DISTINCT FROM :empemail
              AND emplinkedin IS NOT DISTINCT FROM :emplinkedin
            LIMIT 1
        """), {
            'empemail': row['empemail'],
            'emplinkedin': row['emplinkedin']
        }).fetchone()

        if result:
            results.append({'index': row['index'], 'contact_id': result[0]})

    df_matched = pd.DataFrame(results)

    # ✅ Safe merge or fallback
    if df_matched.empty:
        df_contacts['contact_id'] = None
    else:
        df_contacts = df_contacts.merge(df_matched, on='index', how='left')

    return df_contacts


def compare_contacts_to_db(df_contacts: pd.DataFrame, db_session) -> pd.DataFrame:
    """
    Compares contact rows with fact_contacts. Returns df with a 'status' column:
    'Insert', 'Update', or 'No Update'.
    """
    from sqlalchemy import text

    compare_fields = [
        'name', 'firstname', 'lastname', 'empemail', 'emplinkedin',
        'emailstatus_id', 'jobtitle_id', 'company_id',
        'address_id', 'city_id', 'state_id', 'postalcode_id', 'country_id'
    ]

    df_result = df_contacts.copy()
    df_result['id'] = df_result['contact_id']

    status_list = []

    for _, row in df_result.iterrows():
        contact_id = row.get('contact_id')

        if pd.isna(contact_id):
            status_list.append("Insert")
            continue

        query = text(f"""
            SELECT {', '.join(compare_fields)}
            FROM fact_contacts
            WHERE id = :contact_id
        """)
        result = db_session.execute(query, {'contact_id': int(contact_id)}).fetchone()

        if not result:
            status_list.append("Insert")
            continue

        db_row = dict(zip(compare_fields, result))

        is_different = False
        for field in compare_fields:
            df_val = row.get(field)
            db_val = db_row.get(field)

            # Normalize empty strings and NaNs
            if isinstance(df_val, str) and df_val.strip() == '':
                df_val = None
            if isinstance(db_val, str) and db_val.strip() == '':
                db_val = None
            if pd.isna(df_val): df_val = None
            if pd.isna(db_val): db_val = None

            if df_val != db_val:
                is_different = True
                break

        status_list.append("Update" if is_different else "No Update")

    df_result['status'] = status_list
    return df_result

def upsert_contacts(df_contacts: pd.DataFrame, db_session) -> tuple[pd.DataFrame, list[int]]:
    """
    Updates or inserts contacts into fact_contacts.
    Returns:
        - Updated df_contacts with contact_id and id fields
        - List of changed contact IDs (inserted + updated)
    """
    from sqlalchemy import text

    df_result = df_contacts.copy()
    df_result['id'] = df_result['contact_id']

    insert_fields = [
        'name', 'firstname', 'lastname', 'empemail', 'emplinkedin',
        'emailstatus_id', 'jobtitle_id', 'company_id',
        'address_id', 'city_id', 'state_id', 'postalcode_id', 'country_id'
    ]

    changed_ids = []

    # 1. Update rows
    df_update = df_result[df_result['status'] == 'Update']
    for _, row in df_update.iterrows():
        set_clause = ', '.join([f"{col} = :{col}" for col in insert_fields])
        query = text(f"""
            UPDATE fact_contacts
            SET {set_clause}
            WHERE id = :id
        """)
        params = {field: row[field] if pd.notna(row[field]) else None for field in insert_fields}
        params['id'] = int(row['id'])
        db_session.execute(query, params)
        changed_ids.append(int(row['id']))  # track updated ID

    # 2. Insert new rows
    df_insert = df_result[df_result['status'] == 'Insert']
    new_ids = []

    for i, row in df_insert.iterrows():
        insert_query = text(f"""
            INSERT INTO fact_contacts (
                {', '.join(insert_fields)}
            ) VALUES (
                {', '.join([f":{field}" for field in insert_fields])}
            ) RETURNING id
        """)
        params = {field: row[field] if pd.notna(row[field]) else None for field in insert_fields}
        result = db_session.execute(insert_query, params)
        new_id = result.scalar()
        new_ids.append((i, new_id))
        changed_ids.append(new_id)  # track inserted ID

    db_session.commit()

    for idx, contact_id in new_ids:
        df_result.at[idx, 'contact_id'] = contact_id
        df_result.at[idx, 'id'] = contact_id

    print(f"✅ {len(df_update)} contacts updated, {len(new_ids)} contacts inserted into fact_contacts.")
    return df_result, changed_ids

def replace_blank_with_unknown(series: pd.Series) -> pd.Series:
    return (
        series.replace([None, float('nan'), pd.NA], '')  # ensure real nulls handled
            .astype(str)
            .str.strip()
            .replace(['', 'nan', 'NaN', 'None'], 'Unknown')
    )

def replace_blank_with_zero(series: pd.Series) -> pd.Series:
    return series.fillna('').astype(str).str.strip().replace('', '0').astype(int)

def process_uploaded_campaign_file(uploaded_file) -> dict:
    """
    Streamlit-compatible processing function for uploaded QA campaign data.
    Accepts a file-like object (CSV or Excel) and processes contacts and companies.
    Returns: dict with total, inserted, and updated contact record counts.
    """
    import pandas as pd
    from app_backend.database import get_db
    from functions import (
        load_new_data,
        safe_exit,
        replace_nan_with_empty_string,
        convert_qa_disposition,
        convert_zipcode_to_string,
        truncate_linkedin_fields_with_log,
        clean_urls,
        enrich_and_merge_dim,
        get_company_ids,
        compare_companies_to_db,
        upsert_companies,
        get_contact_ids,
        compare_contacts_to_db,
        upsert_contacts
    )

    # Load file from uploaded input
    df = pd.read_excel(uploaded_file) if uploaded_file.name.endswith("xlsx") else pd.read_csv(uploaded_file)

    required_columns = [
        'firstname', 'lastname', 'emplinkedin', 'empemail', 'jobtitle',
        'qa_disposition', 'comp_name', 'comp_domain', 'comp_phone', 'comp_linkedin',
        'annrev', 'empsize', 'comp_street', 'comp_city', 'comp_country',
        'comp_state', 'comp_zipcode', 'comp_industry'
    ]

    if df is None:
        safe_exit("❌ Import stopped. The file could not be loaded. Please check the file format and content.")
    if df is not None and not df.empty:
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            safe_exit(f"❌ Import stopped. Missing required columns: {', '.join(missing_columns)}")

        df = df[~(df['empemail'].astype(str).str.strip() == "")]
        df.reset_index(drop=True, inplace=True)
        df['index'] = df.index + 1
        
        df['name'] = df['firstname'].fillna('') + ' ' + df['lastname'].fillna('')
        
        for col in ['comp_street', 'comp_city', 'comp_state', 'comp_country', 'comp_zipcode', 'comp_industry','jobtitle' ]:
            df[col] = replace_blank_with_unknown(df[col])
        
        df['annrev'] = replace_blank_with_zero(df['annrev'])
        df['empsize'] = replace_blank_with_zero(df['empsize'])
        df['comp_phone'] = df['comp_phone'].fillna('').astype(str).str.strip()
        
        #df = replace_nan_with_empty_string(df, ['emplinkedin', 'jobtitle', 'comp_name', 'comp_domain',  'comp_linkedin'])
        df = convert_qa_disposition(df)
        for col in df.select_dtypes(include='object').columns:
            df[col] = df[col].astype('string')
        #df = convert_zipcode_to_string(df)
        df = truncate_linkedin_fields_with_log(df, ['emplinkedin', 'comp_linkedin'])
        df = clean_urls(df, ['emplinkedin', 'comp_linkedin', 'comp_domain'])

        columns_dims = {
            'jobtitle': 'dim_jobtitles',
            'comp_street': 'dim_addresses',
            'comp_country': 'dim_countries',
            'comp_zipcode': 'dim_postalcodes',
            'comp_industry': 'dim_industries',
            'comp_city': 'dim_cities',
            'comp_state': 'dim_states'
        }

        with next(get_db()) as db:
            df = enrich_and_merge_dim(df, 'comp_country', 'dim_countries', db)
            df = enrich_and_merge_dim(df, 'comp_street', 'dim_addresses', db)
            df = enrich_and_merge_dim(df, 'comp_zipcode', 'dim_postalcodes', db)
            df = enrich_and_merge_dim(df, 'comp_industry', 'dim_industries', db)
            df = enrich_and_merge_dim(df, 'comp_city', 'dim_cities', db)
            df = enrich_and_merge_dim(df, 'comp_state', 'dim_states', db)
            df = enrich_and_merge_dim(df, 'jobtitle', 'dim_jobtitles', db)

        company_columns = {
            'index': 'index',
            'comp_name': 'name',
            'comp_domain': 'comp_domain',
            'comp_linkedin': 'comp_linkedin',
            'comp_phone': 'comp_phone',
            'annrev': 'annrev',
            'empsize': 'empsize',
            'comp_street_id': 'address_id',
            'comp_country_id': 'country_id',
            'comp_zipcode_id': 'postalcode_id',
            'comp_city_id': 'city_id',
            'comp_state_id': 'state_id',
            'comp_industry_id': 'industry_id'
        }

        df_companies = df[list(company_columns.keys())].rename(columns=company_columns)
        for col in ['name', 'comp_domain', 'comp_linkedin']:
            df_companies[col] = df_companies[col].replace('', None)

        with next(get_db()) as db:
            df_companies = get_company_ids(df_companies, db)
            df_companies = compare_companies_to_db(df_companies, db)
            df_companies = upsert_companies(df_companies, db)

        df = df.merge(df_companies[['index', 'company_id']], on='index', how='left')

        df_contacts = df[[
            'index', 'name', 'firstname', 'lastname',
            'emplinkedin', 'empemail', 'qa_disposition',
            'jobtitle_id', 'company_id'
        ]].copy()

        df_contacts = df_contacts.rename(columns={'qa_disposition': 'emailstatus_id'})
        df_contacts['address_id'] = 999999
        df_contacts['city_id'] = 999999
        df_contacts['state_id'] = 999999
        df_contacts['postalcode_id'] = 999999
        df_contacts['country_id'] = 999999

        with next(get_db()) as db:
            # Get contact id's from the database based on matching email and linkedin values
            df_contacts = get_contact_ids(df_contacts, db)
            # Compare with existing records in the database and return df_result
            df_contacts = compare_contacts_to_db(df_contacts, db)
            # Keep only contacts originally in the uploaded file
            # original_indexes = df['index'].tolist()
            # df_contacts = df_contacts[df_contacts['index'].isin(original_indexes)].copy()
            df_contacts, changed_contact_ids = upsert_contacts(df_contacts, db)
            changed_contact_ids = list(set(changed_contact_ids))
            print("📊 Changed contact IDs:", changed_contact_ids)
            from functions import update_cached_contacts
            update_cached_contacts(changed_contact_ids)
            st.cache_data.clear()
            print("🧹 Streamlit cache cleared after DB update.")

        inserted = len(df_contacts[df_contacts['status'] == 'Insert'])
        updated = len(df_contacts[df_contacts['status'] == 'Update'])
        total = len(df_contacts)

        return {
            'total': total,
            'inserted': inserted,
            'updated': updated
        }
    else:
        return {
            'total': 0,
            'inserted': 0,
            'updated': 0
        }

@st.cache_data
def get_filter_options_from_cache():
    """
    Load distinct values for all 10 filters from cached_filters_contacts_data (for stable filter dropdowns).
    Returns a dictionary with all available options.
    """
    query = """
        SELECT DISTINCT country, compstate, city, companyname, industry,
                        empsize, annrev, jobtitle, managementlevel, emailstatus
        FROM cached_filters_contacts_data
    """
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)

    return {
        "country": sorted(df["country"].dropna().unique().tolist()),
        "compstate": sorted(df["compstate"].dropna().unique().tolist()),
        "city": sorted(df["city"].dropna().unique().tolist()),
        "companyname": sorted(df["companyname"].dropna().unique().tolist()),
        "industry": sorted(df["industry"].dropna().unique().tolist()),
        "empsize": sorted(df["empsize"].dropna().unique().tolist()),
        "annrev": sorted(df["annrev"].dropna().unique().tolist()),
        "jobtitle": sorted(df["jobtitle"].dropna().unique().tolist()),
        "managementlevel": sorted(df["managementlevel"].dropna().unique().tolist()),
        "emailstatus": sorted(df["emailstatus"].dropna().unique().tolist())
    }
